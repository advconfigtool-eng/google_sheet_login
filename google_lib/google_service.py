import os
import io
from openpyxl import load_workbook
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload


SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Credentials folder inside project
CREDENTIALS_DIR = os.path.join(BASE_DIR, "credentials")

# Paths
CREDENTIALS_FILE_PATH = os.path.join(CREDENTIALS_DIR, "google_credentials.json")


class GoogleService:
    def __init__(self):
        # Build credentials from the service account file
        creds = service_account.Credentials.from_service_account_file(
            CREDENTIALS_FILE_PATH,
            scopes=SCOPES
        )

        self.creds = creds

    # --- Sheets ---
    def read_sheet(self, spreadsheet_id, range_name):
        service = build("sheets", "v4", credentials=self.creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id, range=range_name
        ).execute()
        return result.get("values", [])

    def append_sheet(self, spreadsheet_id, range_name, values):
        service = build("sheets", "v4", credentials=self.creds)
        body = {
            "values": values
        }
        result = service.spreadsheets().values().append(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            valueInputOption="RAW",   # or "USER_ENTERED"
            insertDataOption="INSERT_ROWS",
            body=body
        ).execute()
        return result

    def write_sheet(self, spreadsheet_id, range_name, values):
        service = build("sheets", "v4", credentials=self.creds)
        body = {"values": values}
        return service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            valueInputOption="RAW",
            body=body
        ).execute()

    def clear_range(self, spreadsheet_id, range_name):
        """Clear values in a given range of Google Sheet"""
        try:
            service = build("sheets", "v4", credentials=self.creds)
            request = service.spreadsheets().values().clear(
                spreadsheetId=spreadsheet_id,
                range=range_name,
                body={}
            )
            response = request.execute()
            return response
        except Exception as e:
            print(f"❌ Error clearing range: {e}")
            return None

    # --- Drive ---
    def download_file(self, file_id, dest_path):
        service = build("drive", "v3", credentials=self.creds)
        request = service.files().get_media(fileId=file_id)
        with io.FileIO(dest_path, "wb") as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
                print(f"Download {int(status.progress() * 100)}%.")
        return dest_path

    def get_file_name(self, file_id):
        service = build("drive", "v3", credentials=self.creds)
        file = service.files().get(fileId=file_id, fields="name").execute()
        return file.get("name")

    def upload_file(self, file_path, mime_type="application/octet-stream", parent_folder_id=None):
        service = build("drive", "v3", credentials=self.creds)
        metadata = {"name": os.path.basename(file_path)}
        if parent_folder_id:
            metadata["parents"] = [parent_folder_id]

        media = MediaFileUpload(file_path, mimetype=mime_type, resumable=True)
        file = service.files().create(
            body=metadata, media_body=media, fields="id"
        ).execute()
        return file.get("id")

    def list_files_in_folder(self, folder_id, query=None):
        result = {
            "is_success": False,
            "err_msg": "",
            "files": []  # array of {id, name, mimeType}
        }

        try:
            service = build("drive", "v3", credentials=self.creds)
            formatted_query = (
                f"'{folder_id}' in parents"
            )
            if query:
                formatted_query += f" and ({query})"
            response = service.files().list(
                q=query,
                fields="files(id, name, mimeType)"
            ).execute()

            result["files"] = response.get("files", [])
            result["is_success"] = True

        except Exception as e:
            result["err_msg"] = f"Error listing files in folder: {e}"

        return result

    def list_latest_files_in_folder(self, folder_id, query=None):
        result = {
            "is_success": False,
            "err_msg": "",
            "files": []  # array of {id, name, mimeType, modifiedTime}
        }

        try:
            service = build("drive", "v3", credentials=self.creds)

            # Always constrain search to the folder
            formatted_query = f"'{folder_id}' in parents"
            if query:
                formatted_query += f" and ({query})"

            response = service.files().list(
                q=formatted_query,
                fields="files(id, name, mimeType, modifiedTime)",
                orderBy="modifiedTime desc"
            ).execute()

            files = response.get("files", [])

            # Keep only the latest per name
            latest_by_name = {}
            for f in files:
                name = f["name"]
                if name not in latest_by_name:
                    latest_by_name[name] = f  # first occurrence = newest due to orderBy

            result["files"] = list(latest_by_name.values())
            result["is_success"] = True

        except Exception as e:
            result["err_msg"] = f"Error listing files in folder: {e}"

        return result

    def read_xlsx_file(self, file_id, sheet_name_list=None):
        """
        Download an Excel file from Google Drive and parse sheets into dict
        :param file_id: Google Drive file ID
        :param sheet_name_list: list of sheet names to read (default = all)
        :return: dict with keys {is_success, err_msg, file_name, file_content}
        """
        result = {
            "is_success": False,
            "err_msg": "",
            "file_name": "",
            "file_content": None  # {Sheet1: [[...], ...]}
        }

        try:
            # --- Get file metadata (name)
            service = build("drive", "v3", credentials=self.creds)
            metadata = service.files().get(fileId=file_id, fields="name").execute()
            result["file_name"] = metadata["name"]

            # --- Download file content into memory
            request = service.files().get_media(fileId=file_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()

            fh.seek(0)

            # --- Load workbook from memory
            wb = load_workbook(filename=fh, data_only=True)

            # --- Extract sheets
            sheets_data = {}
            if sheet_name_list:
                for sheet_name in sheet_name_list:
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        sheets_data[sheet_name] = [
                            list(row)  # row is already a tuple of values
                            for row in ws.iter_rows(values_only=True)
                        ]
            else:
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    sheets_data[sheet_name] = [
                        list(row)  # row is already a tuple of values
                        for row in ws.iter_rows(values_only=True)
                    ]

            result["file_content"] = sheets_data
            result["is_success"] = True

        except Exception as e:
            result["err_msg"] = f"Error reading XLSX file from Google Drive: {e}"

        return result
