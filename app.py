from flask import Flask, jsonify, request, render_template, send_file
from openpyxl import load_workbook
from google_lib.google_service import GoogleService
import time
import os
import datetime
import shutil
from openpyxl.worksheet.datavalidation import DataValidation
from dotenv import load_dotenv


load_dotenv(dotenv_path="/home/repo/google_sheet_login/.env")
# load_dotenv(dotenv_path=".env")


app = Flask(__name__)
gs = GoogleService()

EXCEL_FOLDER_GOOGLE_DRIVE_ID = os.getenv("EXCEL_FOLDER_GOOGLE_DRIVE_ID")
GOOGLE_SHEET_LOGIN_SHEET_ID = os.getenv("GOOGLE_SHEET_LOGIN_SHEET_ID")
FILLING_SHEET_NAME = os.getenv("FILLING_SHEET_NAME", "Fillings")
FILLING_DATA_SHEET_NAME = os.getenv("FILLING_DATA_SHEET_NAME", "FillingsData")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
GENERATED_FOLDER = os.path.join(BASE_DIR, "generated")
EXCEL_MASTER_FILE_ID = os.getenv("MASTER_EXCEL_FILE_ID")
EXCEL_TEMPLATE_FOLDER = os.path.join(BASE_DIR, "excel_templates")

os.makedirs(GENERATED_FOLDER, exist_ok=True)


@app.route("/")
def index():
    return render_template("index.html", filling_options=fetch_filling_options())


@app.route("/api/sync_filling_data", methods=["POST"])
def sync_filling_data():
    fillings_sheet = []
    fillings_data_sheet = []
    option_list = []
    try:
        # 1. List Excel files in folder
        list_file_result = gs.list_files_in_folder(EXCEL_FOLDER_GOOGLE_DRIVE_ID, "mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and trashed=false")
        file_list = list_file_result.get("files", [])

        # 2. Iterate over files
        for file_info in file_list:
            read_file_result = gs.read_xlsx_file(file_info["id"], ["Fillings", "FillingsData"])
            if not read_file_result["is_success"]:
                continue

            file_data = read_file_result["file_content"]

            # Fillings sheet rows
            for row in file_data.get("Fillings", [])[1:]:
                fillings_sheet.append(row)
                option_list.append(row[1])

            # FillingsData sheet rows
            for row in file_data.get("FillingsData", [])[1:]:
                # Map row to expected fields
                if row and len(row) >= 6:
                    fillings_data_sheet.append([
                        row[1],  # Filling Name
                        row[2],  # System Type
                        row[3],  # Module
                        row[4],  # Suffix
                        row[5],  # MaxModules
                    ])

        # 3. Clear old ranges
        gs.clear_range(GOOGLE_SHEET_LOGIN_SHEET_ID, f"'{FILLING_SHEET_NAME}'!A2:G")
        time.sleep(0.1)
        gs.clear_range(GOOGLE_SHEET_LOGIN_SHEET_ID, f"'{FILLING_DATA_SHEET_NAME}'!A2:F")
        time.sleep(0.1)

        # 4. Write new values
        if fillings_sheet:
            gs.write_sheet(GOOGLE_SHEET_LOGIN_SHEET_ID, f"{FILLING_SHEET_NAME}!A2", fillings_sheet)
            time.sleep(0.1)
        if fillings_data_sheet:
            gs.write_sheet(GOOGLE_SHEET_LOGIN_SHEET_ID, f"{FILLING_DATA_SHEET_NAME}!B2", fillings_data_sheet)
            time.sleep(0.1)

        option_list = list(set(option_list))
        current_option_list = []
        filling_order_data = gs.read_sheet(os.getenv("GOOGLE_SHEET_LOGIN_SHEET_ID"), "FillingsOrder!A2:A")
        for row in filling_order_data:
            if row[0]:
                current_option_list.append(row[0])

        new_option_list = []
        for option in current_option_list:
            if option in option_list:
                new_option_list.append([option])
        for option in option_list:
            if option not in current_option_list:
                new_option_list.append([option])
        gs.clear_range(GOOGLE_SHEET_LOGIN_SHEET_ID, f"FillingsOrder!A2:A")
        time.sleep(0.1)
        gs.write_sheet(GOOGLE_SHEET_LOGIN_SHEET_ID, f"FillingsOrder!A2", new_option_list)
        time.sleep(0.1)

        # 5. Return JSON
        return jsonify({
            "fillings_sheet": fillings_sheet,
            "fillings_data_sheet": fillings_data_sheet
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/get_filling_options", methods=["GET"])
def get_filling_options():
    return jsonify(fetch_filling_options())


def fetch_filling_options():
    try:
        sheet_id = os.getenv("GOOGLE_SHEET_LOGIN_SHEET_ID")
        read_sheet_data = gs.read_sheet(sheet_id, f"FillingsOrder!A2:A")
        # Extract values safely
        option_list = [
            row[0] for row in read_sheet_data if row and row[0]
        ]
        # Deduplicate while preserving order
        option_list = list(dict.fromkeys(option_list))
        return option_list

    except Exception as e:
        return jsonify({"error": f"Failed to fetch filling options: {str(e)}"}), 500


@app.route("/api/download_template_file", methods=["POST"])
def download_template_file():
    try:
        folder_id = os.getenv("EXCEL_FOLDER_GOOGLE_DRIVE_ID")
        # 1. List Excel files in Google Drive folder
        list_file_result = gs.list_files_in_folder(
            folder_id,
            "(mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' \
            or mimeType='application/vnd.ms-excel.sheet.macroEnabled.12') and trashed=false"
        )

        file_list = list_file_result.get("files", [])

        # 2. Download each file into TEMPLATE_FOLDER
        for file_info in file_list:
            app.logger.info(file_info)
            file_path = os.path.join(EXCEL_TEMPLATE_FOLDER, file_info["name"])
            gs.download_file(file_info["id"], file_path)

        return jsonify({"is_success": True})

    except Exception as e:
        app.logger.error(f"Error in download_template_file: {e}")
        return jsonify({"is_success": False, "err_msg": str(e)}), 500


def validate_input(filling_options, loading_codes):
    result = {"is_success": False, "err_msg": ""}

    # split loading codes into list
    loading_codes_list = [x.strip() for x in loading_codes.split(",") if x.strip()]

    if not filling_options:
        result["err_msg"] = "Mandatory field Filling Options is empty"
        return result

    # --- read sheets
    read_filling_result = gs.read_sheet(
        os.getenv("GOOGLE_SHEET_LOGIN_SHEET_ID"),
        f"'{os.getenv('FILLING_SHEET_NAME')}'!A1:G"
    )
    read_filling_data_result = gs.read_sheet(
        os.getenv("GOOGLE_SHEET_LOGIN_SHEET_ID"),
        f"'{os.getenv('FILLING_DATA_SHEET_NAME')}'!B1:F"
    )

    # --- build filling_data_dict
    filling_data_dict = {}
    filling_data_header = read_filling_data_result[0]
    for row_data in read_filling_data_result[1:]:
        row_dict = get_row_dict(row_data, filling_data_header)
        if row_dict.get("Filling Name"):
            filling_data_dict.setdefault(row_dict["Filling Name"], [])
            filling_data_dict[row_dict["Filling Name"]].append(row_data)

    # --- build filling_info_dict
    filling_info_dict = {}
    filling_header = read_filling_result[0]
    for row_data in read_filling_result[1:]:
        row_dict = get_row_dict(row_data, filling_header)
        filling_visible_name = row_dict.get("Visible Name")
        filling_name = row_dict.get("Filling Name")

        if filling_visible_name:
            filling_info_dict.setdefault(filling_visible_name, {})

            dependencies_splited = row_dict.get("Dependencies", "").split(",")
            dependencies_list = [d.strip() for d in dependencies_splited if d.strip()]
            dependencies_list.append(row_dict.get("SpreadSheet Name"))

            filling_info_dict[filling_visible_name][filling_name] = {
                "row_data": row_data,
                "LC_code": row_dict.get("Loading Code", ""),
                "sheet_name": row_dict.get("SpreadSheet Name"),
                "sheet_id": row_dict.get("SpreadSheet ID"),
                "filling_data": filling_data_dict.get(filling_name, []),
                "dependencies": dependencies_list,
            }

    # --- validate requested filling_options
    validated_filling_dict = {}
    for filling_option in filling_options:
        if filling_option not in filling_info_dict:
            result["err_msg"] = f"Filling options: {filling_option} is not available. Please contact developer"
            return result

        validated_filling_name_list = []
        filling_code_dict = filling_info_dict[filling_option]

        for filling_name, details in filling_code_dict.items():
            expected_loading_code = details.get("LC_code", "")
            if expected_loading_code in loading_codes_list or expected_loading_code == "":
                validated_filling_name_list.append(filling_name)
                loading_codes_list = [val for val in loading_codes_list if val != expected_loading_code]

        if not validated_filling_name_list:
            result["err_msg"] = f"Option code not found for filling: {filling_option}."
            return result

        validated_filling_dict[filling_option] = {
            filling_name: filling_info_dict[filling_option][filling_name]
            for filling_name in validated_filling_name_list
        }

    result["is_success"] = True
    result["validated_filling_dict"] = validated_filling_dict
    return result

@app.route("/api/generate_excel_files", methods=["POST"])
def generate_excel_files():
    try:
        body = request.get_json()
        filling_options = body.get("filling_options", [])
        loading_codes = body.get("loading_codes", "")

        result = {"is_success": False, "err_msg": "", "file_content": ""}

        # 🔹 Validate input
        validate_result = validate_input(filling_options, loading_codes)
        if not validate_result["is_success"]:
            return jsonify({"is_success": False, "err_msg": validate_result["err_msg"]})

        validated_filling_dict = validate_result["validated_filling_dict"]

        filling = []
        filling_data = []
        dependencies = []

        # 🔹 Collect data
        for filling_option, names in validated_filling_dict.items():
            for filling_name, details in names.items():
                filling.append(details["row_data"])
                filling_data.extend(details["filling_data"])
                dependencies.extend(details["dependencies"])

        dependencies = list(dict.fromkeys(dependencies))  # dedupe

        # 🔹 Create timestamped file name
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        master_file_name = gs.get_file_name(EXCEL_MASTER_FILE_ID)
        master_file_path = os.path.join(EXCEL_TEMPLATE_FOLDER, master_file_name)
        file_name = f"{master_file_name}_{'_'.join(filling_options)}_{timestamp}.xlsm"
        copy_path = os.path.join(GENERATED_FOLDER, file_name)

        # 🔹 Copy master file
        shutil.copy(master_file_path, copy_path)

        # 🔹 Load workbook
        wb = load_workbook(copy_path, keep_vba=True)

        # 🔹 Write Fillings
        if "Fillings" in wb.sheetnames and filling:
            ws = wb["Fillings"]
            for r, row in enumerate(filling, start=2):
                for c, val in enumerate(row, start=1):
                    ws.cell(row=r, column=c, value=val)

        # 🔹 Write FillingsData
        if "FillingsData" in wb.sheetnames and filling_data:
            ws = wb["FillingsData"]
            for r, row in enumerate(filling_data, start=2):  # write starting at row 2
                for c, val in enumerate(row, start=2):
                    if isinstance(val, str) and val.strip().replace(".", "", 1).isdigit():
                        # Convert to int or float depending on presence of "."
                        if "." in val:
                            val = float(val)
                        else:
                            val = int(val)
                    ws.cell(row=r, column=c, value=val)
        rebuild_data_validation(wb)

        # 🔹 Copy dependency sheets (if exist)
        for dep in dependencies:
            dep_file = os.path.join(EXCEL_TEMPLATE_FOLDER, f"{dep}.xlsx")
            if os.path.exists(dep_file):
                dep_wb = load_workbook(dep_file)
                for sheet_name in dep_wb.sheetnames:
                    if sheet_name not in ("Fillings", "FillingsData") and "." in sheet_name:
                        dep_ws = dep_wb[sheet_name]
                        copied_ws = copy_sheet_values(dep_ws, wb, f"{sheet_name}.{dep}")
                        copied_ws.sheet_state = "hidden"

        # 🔹 Save workbook
        wb.save(copy_path)
        wb.close()

        return send_file(
            copy_path,
            as_attachment=True,
            download_name=file_name,
            mimetype="application/vnd.ms-excel.sheet.macroEnabled.12"
        )

    except Exception as e:
        app.logger.error(f"❌ Excel generation failed: {e}")
        return jsonify({"is_success": False, "err_msg": str(e)}), 500


def get_row_dict(row_data, header):
    row_dict = {}
    for index, col_name in enumerate(header):
        if str(col_name):  # only if header cell has value
            if index < len(row_data):
                row_dict[col_name] = row_data[index]
            else:
                row_dict[col_name] = ""
    return row_dict


def copy_sheet_values(source_ws, target_wb, new_title):
    target_ws = target_wb.create_sheet(title=new_title)
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws[cell.coordinate]
            new_cell.value = cell.value
    return target_ws


def rebuild_data_validation(wb):
    configurator_sheet = wb["Configurator"]

    # Example: Dropdown in D7 from Fillings!A2:A50
    dv_main = DataValidation(
        type="list",
        formula1="=Fillings!$A$2:$A$50",
        allow_blank=True,
        showDropDown=False
    )
    configurator_sheet.add_data_validation(dv_main)
    dv_main.add("D7:I7")

    # Row-based dependent dropdowns
    start_row = 15
    end_row = 50

    for r in range(start_row, end_row + 1):
        # IMPORTANT: formula must be quoted as string for Excel
        formula = f'=$AD${r}:$BJ${r}'

        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False
        )
        configurator_sheet.add_data_validation(dv)
        dv.add(configurator_sheet[f"F{r}"])

    dv_head = DataValidation(
        type="list",
        formula1='=TestHeads!$K$2:$K$50',
        allow_blank=True,
        showDropDown=False
    )
    configurator_sheet.add_data_validation(dv_head)
    dv_head.add(configurator_sheet["F12"])


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)

